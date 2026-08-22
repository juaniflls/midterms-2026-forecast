from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def normalize(value):
    if hasattr(value, "text") or hasattr(value, "ref"):
        return {
            "type": type(value).__name__,
            "text": getattr(value, "text", None),
            "ref": getattr(value, "ref", None),
            "ca": getattr(value, "ca", None),
        }
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return {"type": type(value).__name__, "repr": str(value)}


def workbook_signature(path: Path):
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheets = []
    for worksheet in workbook.worksheets:
        cells = {}
        formula_count = 0
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                if cell.value is None:
                    continue
                value = normalize(cell.value)
                cells[cell.coordinate] = {"value": value, "data_type": cell.data_type}
                if cell.data_type == "f" or type(cell.value).__name__.endswith("Formula"):
                    formula_count += 1
        sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "merged_ranges": sorted(str(item) for item in worksheet.merged_cells.ranges),
                "cells": cells,
                "formula_count": formula_count,
            }
        )
    defined_names = sorted(
        (item.name, item.attr_text, item.localSheetId, item.hidden)
        for item in workbook.defined_names.values()
    )
    return {"sheets": sheets, "defined_names": defined_names}


def main() -> None:
    before = Path(sys.argv[1])
    after = Path(sys.argv[2])
    output = Path(sys.argv[3])
    left = workbook_signature(before)
    right = workbook_signature(after)
    result = {
        "source": str(before),
        "styled": str(after),
        "sheet_order_equal": [s["name"] for s in left["sheets"]]
        == [s["name"] for s in right["sheets"]],
        "defined_names_equal": left["defined_names"] == right["defined_names"],
        "sheets": [],
    }
    right_by_name = {sheet["name"]: sheet for sheet in right["sheets"]}
    all_equal = result["sheet_order_equal"] and result["defined_names_equal"]
    for source_sheet in left["sheets"]:
        styled_sheet = right_by_name.get(source_sheet["name"])
        if styled_sheet is None:
            result["sheets"].append({"name": source_sheet["name"], "missing": True})
            all_equal = False
            continue
        source_cells = source_sheet["cells"]
        styled_cells = styled_sheet["cells"]
        differing = []
        for coordinate in sorted(set(source_cells) | set(styled_cells)):
            if source_cells.get(coordinate) != styled_cells.get(coordinate):
                differing.append(
                    {
                        "coordinate": coordinate,
                        "source": source_cells.get(coordinate),
                        "styled": styled_cells.get(coordinate),
                    }
                )
                if len(differing) >= 25:
                    break
        sheet_result = {
            "name": source_sheet["name"],
            "values_and_formulas_equal": not differing,
            "merged_ranges_equal": source_sheet["merged_ranges"] == styled_sheet["merged_ranges"],
            "source_nonempty_cells": len(source_cells),
            "styled_nonempty_cells": len(styled_cells),
            "source_formula_count": source_sheet["formula_count"],
            "styled_formula_count": styled_sheet["formula_count"],
            "differences_sample": differing,
        }
        if not sheet_result["values_and_formulas_equal"] or not sheet_result["merged_ranges_equal"]:
            all_equal = False
        result["sheets"].append(sheet_result)
    result["logic_and_structure_equal"] = all_equal
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "logic_and_structure_equal": all_equal,
        "sheet_order_equal": result["sheet_order_equal"],
        "defined_names_equal": result["defined_names_equal"],
        "sheet_mismatches": [
            sheet["name"] for sheet in result["sheets"]
            if not sheet.get("values_and_formulas_equal", False)
            or not sheet.get("merged_ranges_equal", False)
        ],
    }, ensure_ascii=False))
    if not all_equal:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
